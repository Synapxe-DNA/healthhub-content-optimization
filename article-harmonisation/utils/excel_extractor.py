import warnings

import pandas as pd
from openpyxl import load_workbook
from states.definitions import OptimisationFlags

# Filters out warnings
warnings.simplefilter(action="ignore", category=UserWarning)


def return_optimisation_flags(article, rewriting_process: str):
    """
    Returns the user flags for article rewriting. Article harmonisation will return all flags as True while article optimisation will be dependent on the user flags in the User Annotation Excel File

    Args:
        article (DataFrame): A DataFrame consisting of articles from the User Annotation Excel file.
        rewriting_process (str): A String indicating which rewriting process. It can either be "optimisation" or "harmonisation"

    Returns:
        flags (OptimisationFlags): A TypedDict with keys for each optimisation flag. The flags will either be True or False, depending on the user input in the User Annotation Excel file

    """
    # Declaring the optimisation flags
    flags = OptimisationFlags(
        flag_for_content_optimisation=True,
        flag_for_title_optimisation=True,
        flag_for_meta_desc_optimisation=True,
        flag_for_writing_optimisation=True,
    )

    match rewriting_process:
        case "optimisation":
            # Extracting article actions from Excel
            article_actions = article["Action"]

            # Checking for user annotated flags
            if "Send for title optimisation" not in article_actions:
                flags["flag_for_title_optimisation"] = False
            if "Send for meta description optimisation" not in article_actions:
                flags["flag_for_meta_desc_optimisation"] = False
            if "Send for content optimisation" not in article_actions:
                flags["flag_for_writing_optimisation"] = False

            # Checks if the content category is not diseases and conditions, as it will not require content optimisation if it's not.
            if article["content category"] not in ["diseases-and-conditions"]:
                flags["flag_for_content_optimisation"] = False

            return flags
        case "harmonisation":
            if article["content category"] not in [
                "diseases-and-conditions",
                "live-healthy-articles",
            ]:
                flags["flag_for_content_optimisation"] = False
            # Returning the other flags as true as this is a harmonisation flow
            return flags


def store_optimised_outputs(file_path: str, sheet_name: str, article_data):
    """
    Stores the optimised outputs into an Excel file.

    Args:
        file_path (str): A String indicating the file path to the destination Excel file
        sheet_name (str): A String indicating the name of the sheet to store the output in
        article_data (list): A List storing the article data to be stored
    """

    # List containing all column headers for article harmonisation Excel sheet
    article_harmonisation_columns = [
        "Group Number",
        "Subgroup",
        "urls",
        "Group Description",
        "Article ids",
        "Title Option 1",
        "Title Option 2",
        "Title Option 3",
        "Title Chosen",
        "Optional: Title written by user",
        "Meta Description 1",
        "Meta Description 2",
        "Meta Description 3",
        "Meta Description Chosen",
        "Optional: Meta Description written by user",
        "Optimised article content",
        "Article optimisation evaluation summary",
        "User approval of optimised article",
        "Optional: User attached updated article (Y)",
        "Content Edit Status (if any)",
    ]

    # List containing all column headers for article optimisation Excel sheet
    article_optimisation_columns = [
        "article id",
        "url",
        "Overall title flag",
        "long title",
        "irrelevant title",
        "Reasons for irrelevant title",
        "Optimised Title 1",
        "Optimised Title 2",
        "Optimised Title 3",
        "Title Chosen",
        "Optional: Title written by user",
        "Overall meta description flag",
        "meta description not within 70 and 160 characters",
        "irrelevant meta description",
        "Reasons for irrelevant meta description",
        "Optimised Meta Description 1",
        "Optimised Meta Description 2",
        "Optimised Meta Description 3",
        "Meta Description Chosen",
        "Optional: Meta Description written by user",
        "Overall content flag",
        "poor readability",
        "reasons for poor readability",
        "insufficient content",
        "Optimised article content",
        "Article optimisation evaluation summary",
        "User approval of optimised article",
        "Optional: User attached updated article (Y)",
        "Content Edit Status (if any)",
        "Content Filter Flag",
    ]

    # try statement that checks if Excel file already exists
    try:
        # Load the existing workbook
        workbook = load_workbook(file_path)

        # Checks if sheetname already exists in the workbook
        if sheet_name in workbook.sheetnames:
            # Edit the existing sheet
            sheet = workbook[sheet_name]
            print(f"Editing existing sheet: {sheet_name}")
        else:
            # Create a new sheet
            sheet = workbook.create_sheet(title=sheet_name)
            print(f"Creating new sheet: {sheet_name}")

            # Adding the article optimisation column headers to User Annotation (Optimised) sheet
            if sheet_name == "User Annotation (Optimised)":
                sheet.append(article_optimisation_columns)

            # Adding the article harmonisation column headers to User Annotation (Harmonised) sheet
            elif sheet_name == "User Annotation (Harmonised)":
                sheet.append(article_harmonisation_columns)

            # Value Error raised if sheet_name is invalid
            else:
                raise ValueError(f"{sheet_name} is not a valid sheet name!")

        # Adding article data to the sheet
        sheet.append(article_data)

        # Save the workbook
        workbook.save(file_path)
        print(f"Workbook '{file_path}' saved successfully.")

    except FileNotFoundError:
        print(f"Workbook '{file_path}' not found, creating new Excel file")

        # If sheet_name is "User Annotation (Optimised)", creates a list containing the article_optimisation_columns and article_data
        if sheet_name == "User Annotation (Optimised)":
            columns = article_optimisation_columns
        # elif sheet_name is "User Annotation (Harmonisaed)", creates a list containing the article_harmonisation_columns and article_data
        elif sheet_name == "User Annotation (Harmonised)":
            columns = article_harmonisation_columns
        # else raise ValueError
        else:
            raise ValueError(f"{sheet_name} is not a valid sheet name!")

        # Creates a dataframe with first list in data as the column headers and subsequent rows as data
        df = pd.DataFrame([article_data], columns=columns)

        # Converting the dataframe to an Excel sheet
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
        print("Excel file sucessfully created")
